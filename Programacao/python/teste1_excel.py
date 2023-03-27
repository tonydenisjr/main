from openpyxl import Workbook, load_workbook

wb = load_workbook("C:\Programacao\main\Programacao\python\excel_teste1.xlsx")
ws = wb.active
ws['A1'].value = "Laranja"


wb.save("C:\Programacao\main\Programacao\python\excel_teste1.xlsx")






